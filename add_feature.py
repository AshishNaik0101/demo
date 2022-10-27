import os
import xlsxwriter
#import StringIO
from datetime import datetime, timedelta
import shutil
import platform
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.styles import numbers
from openpyxl.utils import FORMULAE
from copy import copy
import time
#Source="Feature.xlsx"
Source="D:\\VW_SWRT_RegressionCheck\\VW_SWRT_RegressionCheck.xlsx"
Protocol=load_workbook(filename = Source)
TA_result=Protocol['SWRT Regression Check (auto)'] 
TA_DOORS=Protocol['Test_case_State_DOORS'] 
FB_result=Protocol['Feature_based_results'] 
start_index_TA=16
end_index_TA=750
start_index_D=2
end_index_D=1130
start_index_F=16
end_index_F=645
Test_list_TA=[]
Test_list_D=[]
Feature_list_D=[]
temp=""
temp1=[]
read=()
Feature_dict_D={}

C_F_P_count=0
C_F_F_count=0
C_F_E_count=0
C_F_N_count=0
count=16
C_num='B1'
C_col_num='B2'
start_index_F=int(FB_result[C_num].value)
C_colum=(str(FB_result[C_col_num].value))

#used for the feature name update in Regression sheet
for i in range(start_index_TA,end_index_TA+1):
    for j in range(start_index_D,end_index_D+1):
        if((str(TA_result[('C'+str(i))].value).strip()) == (str(TA_DOORS[('B'+str(j))].value)).strip()):
            TA_result[('B'+str(i))].value=str(TA_DOORS[('D'+str(j))].value)
            print(str(TA_DOORS[('D'+str(j))].value))

#Seperate each feature and link no of test cases 
"""
Feature_list_D=str(TA_DOORS[('D'+str(46))].value)
print(Feature_list_D)
for i in len(Feature_list_D):
    Feature_dict_D[Feature_list_D[i]]=TA_DOORS[('B'+str(46))].value
"""
for k in range(start_index_D,end_index_D+1):
    temp=(str(TA_DOORS[('D'+str(k))].value)).replace("\n", ",")
    if("," in temp):
        temp1=temp.split(",")
        for l in range(len(temp1)):
            if(temp1[l] in Feature_list_D):
                pass
            else:
                Feature_list_D.append((temp1[l]))
    else:
        if("None" == temp):
            pass
        else:
            if(temp in Feature_list_D):
                pass
            else:
                Feature_list_D.append(temp)
print(Feature_list_D)

print("Excel read")
end_index_F=len(Feature_list_D)+start_index_F
c_date=(str(TA_result[C_colum+'1'].value))
t=0
for t in range(len(Feature_list_D)):
    FB_result[('A'+str(start_index_F+t))].value=str(c_date)
    FB_result[('B'+str(start_index_F+t))].value=Feature_list_D[t]
Protocol.save(Source)
k=0
#column wise update
for i in range(start_index_F,end_index_F):
    print("running.....")
    for j in range(start_index_D,end_index_D):
        if((str(FB_result[('B'+str(i))].value).strip()) in  str(TA_DOORS[('D'+str(j))].value)):
            FB_result[('H'+str(i))].value =str(FB_result[('H'+str(i))].value) +","+str(TA_DOORS[('B'+str(j))].value)
            for k in range(start_index_TA,end_index_TA):
                if(str(TA_DOORS[('B'+str(j))].value) in str(TA_result[('B'+str(k))].value)):
                    if(str(TA_result[(C_colum+str(k))].value).lower() == "failed"):
                        C_F_F_count=C_F_F_count+1
                    elif(str(TA_result[(C_colum+str(k))].value).lower() == "success"):
                        C_F_P_count=C_F_P_count+1
                    elif(str(TA_result[(C_colum+str(k))].value).lower() == "error"):
                        C_F_E_count=C_F_E_count+1
                    else:
                        C_F_N_count=C_F_N_count+1
    FB_result[('D'+str(i))].value=C_F_P_count
    FB_result[('E'+str(i))].value=C_F_F_count
    FB_result[('F'+str(i))].value=C_F_E_count
    FB_result[('G'+str(i))].value=C_F_N_count
    FB_result[('C'+str(i))].value=str(C_F_P_count+C_F_F_count+C_F_E_count+C_F_N_count)
    #count=count+1
    C_F_P_count=0
    C_F_F_count=0
    C_F_E_count=0
    C_F_N_count=0
FB_result[C_num].value=end_index_F
Protocol.save(Source)
Protocol.close()
print("updated")




"""
row wise update
for i in range(start_index_F,end_index_F+1):
    for j in range(start_index_TA,end_index_TA+1):
        if(str(FB_result[('A'+str(i))].value) in  str(TA_result[('B'+str(j))].value)):
            FB_result[('G'+str(i))].value =str(FB_result[('G'+str(i))].value) +","+str(TA_DOORS[('B'+str(j))].value)
            if(str(TA_result[(C_colum+str(j))].value) == "Failed"):
                C_F_F_count=C_F_F_count+1
            if(str(TA_result[(C_colum+str(start_index_TA))].value) == "Passed"):
                C_F_P_count=C_F_P_count+1
            if(str(TA_result[(C_colum+str(start_index_TA))].value) == "Error"):
                C_F_E_count=C_F_E_count+1
            else:
                C_F_N_count=C_F_N_count+1
    FB_result[('C'+str(i))].value=C_F_P_count
    FB_result[('D'+str(i))].value=C_F_F_count
    FB_result[('E'+str(i))].value=C_F_E_count
    FB_result[('F'+str(i))].value=C_F_N_count
    FB_result[('B'+str(i))].value=str(C_F_P_count+C_F_F_count+C_F_E_count+C_F_N_count)
    #count=count+1
    C_F_P_count=0
    C_F_F_count=0
    C_F_E_count=0
    C_F_N_count=0
Protocol.save(Source)
Protocol.close()
print("updated")
"""
"""
for i in range(start_index_TA,end_index_TA+1):
    for j in range(start_index_D,end_index_D+1):
        if(str(FB_result[('A'+str(count))].value) in  str(TA_DOORS[('D'+str(j))].value)):
            FB_result[('G'+str(count))].value =str(FB_result[('G'+str(count))].value) +","+str(TA_DOORS[('B'+str(j))].value)
            print(str(TA_DOORS[('B'+str(j))].value))
    #count=count+1
Protocol.save(Source)
"""