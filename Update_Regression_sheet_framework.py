###########################################START#######################################
import os.path
import xlsxwriter
import StringIO
from datetime import datetime, timedelta
import shutil
import platform
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.styles import numbers
from openpyxl.utils import FORMULAE
from copy import copy
import time
Source="\\\\vt1.vitesco.com\\SMT\did01227\\HEV_PNS_SW\\01_Base_Development\\07_Test\\06_Jenkins"
#Source="\\\\vt1.vitesco.com\\SMT\\did01227\\HEV_PNS_SW\\temp\\Jenkins"
#Regression_sheet="\\\\vt1.vitesco.com\\SMT\\did01227\\HEV_PNS_SW\\temp\\Jenkins\\VW_SWRT_RegressionCheck.xlsx"
Regression_sheet="\\\\vt1.vitesco.com\\SMT\\did01227\\HEV_PNS_SW\\01_Base_Development\\07_Test\\06_Jenkins\\VW_SWRT_RegressionCheck.xlsx"
# Test run .txt files path to read plannes test cases list
EMA="\\\\TSD6977W\\Jenkins\\FLASH\\Test_run.txt"
DIAG="\\\\rbdAWUPW\\Jenkins\\FLASH\\Test_run.txt"
SMF="\\\\ILDL149W\\Jenkins\\FLASH\\Test_run.txt"
GEN="\\\\ildk919W\\Jenkins\\FLASH\\Test_run.txt"
BSW="\\\\ildl150w\\Jenkins\\FLASH\\Test_run.txt"
#Reports=['\\\\ildl150w\\Jenkins\\workspace\\VWH02_0U0_TA\\VWH02_TA_HIL\\Test_report_summary.xlsx','\\\\PTD2002W\\Jenkins\\workspace\\VWH02_0U0_TA\\VWH02_TA_HIL\\Test_report_summary.xlsx','\\\\tsd6977w.vt1.vitesco.com\\Jenkins\\workspace\\VWH02_0U0_TA\\VWH02_TA_HIL\\Test_report_summary.xlsx','\\\\rbdAWUPW\\Jenkins\\workspace\\VWH02_0U0_TA\\VWH02_TA_HIL\\Test_report_summary.xlsx']
EMA_count=[]
DIAG_count=[]
SMF_count=[]
GEN_count=[]
BSW_count=[]
cell_n=0
cell_p=0
T_Test_cases_list=[]
Test_cases_list=[]
Test_results_list=[]
start_index=9
max_testcases=748
start_index_Reg=16
end_index=0
file_name_construct=""
grey_fill = PatternFill(fgColor='A5A5A5',bgColor='A5A5A5',fill_type='solid')
white_smoke_fill=PatternFill(fgColor='EDEDED',bgColor='EDEDED',fill_type='solid')
light_green_fill=PatternFill(fgColor='C6EFCE',bgColor='C6EFCE',fill_type='solid')
light_pink_fill=PatternFill(fgColor='FFC7CE',bgColor='FFC7CE',fill_type='solid')
smoke_yellow_fill=PatternFill(fgColor='FFEB9C',bgColor='FFEB9C',fill_type='solid')
light_yellow_fill=PatternFill(fgColor='FFC000',bgColor='FFC000',fill_type='solid')
date = datetime.now().strftime('%d-%m-%Y')
print(date)
date_format = datetime.now().strftime('%d_%m_%Y')


# update on customized date
date_format='26_10_2022'
date='26-10-2022'
Source=Source+'\\'+date
print(date)

List_protocols = os.listdir(Source)
print(List_protocols)
def Read_Test_run_txt():
    f=open(GEN,"r")
    GEN_count=(f.readlines()) 
    f.close()
    f=open(BSW,"r")
    BSW_count=(f.readlines()) 
    f.close()
    f=open(EMA,"r")
    EMA_count=(f.readlines()) 
    f.close()
    f=open(DIAG,"r")
    DIAG_count=(f.readlines()) 
    f.close()
    f=open(SMF,"r")
    SMF_count=(f.readlines()) 
    f.close()
    T_count=EMA_count+DIAG_count+SMF_count+GEN_count+BSW_count
    print(len(EMA_count))
    print(len(DIAG_count))
    print(len(SMF_count))
    print(len(GEN_count))
    print(len(BSW_count))
    print(len(T_count))
    return (len(T_count),len(EMA_count),len(DIAG_count),len(SMF_count),len(GEN_count),len(BSW_count))

def Find_Last_cellnumber(sheet_name,start_index, max_testcases):
    for cell in range(start_index,max_testcases):
        if((sheet_name['B'+str(cell)].value)!="None"):
            pass
        else:
            break
    return cell
                                                                                          
def Find_last_column(start_date):
    #start_date= "13-01-2022"
    yesterday_date = (datetime.now() - timedelta(1)).strftime('%d-%m-%Y')
    date_format ='%d-%m-%Y'
    a = datetime.strptime(start_date, date_format)
    b = datetime.strptime(yesterday_date, date_format)
    delta = b - a
    print(delta.days)  
    cell= delta.days
    cell_n=int(cell)
    cell_n=cell_n+5
    column_name=""
    MAX = 50
    name = ["\0"] * MAX
    i = 0
    while cell > 0:
        rem = cell % 26
        if rem == 0:
            name[i] = 'Z'
            i += 1
            cell = (cell / 26) - 1
        else:
            name[i] = chr((rem - 1) + ord('A'))
            i += 1
            cell = cell / 26
    name[i] = '\0'
    name = name[::-1]
    column_name="".join(name)
    print name
    column_name= column_name.lstrip('\x00')
    column_name= column_name.rstrip('\x00')
    print column_name
    return column_name,cell_n
Regression=load_workbook(filename = Regression_sheet)
sheet_ranges=Regression['SWRT Regression Check (auto)'] 
Plan_sheet=Regression['Plan'] 
Last_column,cell_n=Find_last_column("13-01-2022")
# pass 13-01-2022 for today update, pass 14-01-2022 for yesterday update

print(Last_column)
sheet_ranges[Last_column+str(1)]=date
sheet_ranges[Last_column+str(2)]="TD5_VWH02_0U0_NIGHTLY - Build #"+str(cell_n)
sheet_ranges[Last_column+str(15)]="Result"
sheet_ranges[Last_column+str(15)].fill=PatternFill(fgColor='DDEBF7',bgColor='DDEBF7',fill_type='solid')
Planned_test_count,EMA_count,DIAG_count,SMF_count,GEN_count,BSW_count=Read_Test_run_txt()
sheet_ranges[Last_column+str(5)]=Planned_test_count
Plan_sheet['A'+str(cell_n-13)]=datetime.now().strftime('%d-%b-%y')
Plan_sheet['B'+str(cell_n-13)]=(GEN_count)
Plan_sheet['C'+str(cell_n-13)]=(BSW_count)
Plan_sheet['D'+str(cell_n-13)]=(EMA_count)
Plan_sheet['E'+str(cell_n-13)]=(DIAG_count)
Plan_sheet['F'+str(cell_n-13)]=(SMF_count)
Plan_sheet['G'+str(cell_n-13)]=Planned_test_count
for f in List_protocols:
    Protocol = load_workbook(filename = Source+"\\"+f)
    sheet = Protocol['Sheet1'] 
    end_index=Find_Last_cellnumber(sheet,start_index,max_testcases)
    print(end_index)
    for i in range(start_index,end_index+1):
        if(str(sheet[('B'+str(i))].value)!="None"):
            Test_cases_list.append(str(sheet[('B'+str(i))].value))
            print(str(sheet[('B'+str(i))].value))
            Test_results_list.append(str(sheet[('E'+str(i))].value)) 
    Protocol.close()
    #print(file_name_construct)
    #if( f == file_name_construct):
    #    print(str(sheet_ranges[('A'+str(18))].value)) 
    #    print(f)
    print(f)
    print(len(Test_results_list))
    print(Test_results_list)
    for j in range(len(Test_cases_list)):
        test_present=0
        for k in range(start_index_Reg,max_testcases):
            file_name_construct=("Test_report_summary_"+(str(sheet_ranges[('A'+str(k))].value).strip()+"_"+date_format+".xlsx"))
            if(str(Test_cases_list[j]).strip() == str(sheet_ranges[('B'+str(k))].value ) ):
                #print(Test_cases_list[j] + "  "+Test_result_list[j])
                if(f==file_name_construct):
                    #print("file matched")
                    if(Test_results_list[j]=="Success"):
                        sheet_ranges[(Last_column+str(k))]="Success"
                        #sheet_ranges[(Last_column+str(k))].fill=greenFill
                    if(Test_results_list[j]=="Failed"):
                        sheet_ranges[(Last_column+str(k))]="Failed"
                        #sheet_ranges[(Last_column+str(k))].fill=redFill
                    if(Test_results_list[j]=="Error"):
                        sheet_ranges[(Last_column+str(k))]="Error"
                        #sheet_ranges[(Last_column+str(k))].fill=yellowFill
                    else:
                        sheet_ranges[(Last_column+str(k))]=Test_results_list[j]
                else :
                    print("Duplicate Test case from Feature team  " +str(sheet_ranges[('A'+str(k))].value)+ " in Regression sheet ")
                    print(Test_cases_list[j])
                test_present=1
        if(test_present==0):
            print(" Test case to be added in Regression sheet")
            print(Test_cases_list[j])
    T_Test_cases_list=T_Test_cases_list+Test_cases_list
    Test_cases_list=[]
    Test_results_list=[]
print(len(T_Test_cases_list))
Plan_sheet['L'+str(cell_n-13)]=len(T_Test_cases_list)
sheet_ranges[Last_column+str(6)]=len(T_Test_cases_list)
time.sleep(5)
sheet_ranges[Last_column+str(7)].number_format =numbers.FORMAT_PERCENTAGE
sheet_ranges[Last_column+str(7)]="=((COUNTIF("+(Last_column+str(start_index_Reg))+":"+(Last_column+str(max_testcases))+","+ '"Success"'+")"+"/"+Last_column+str(5)+"))"
P_Last_column, cell_p=Find_last_column("14-01-2022")
# pass 14-01-2022 for today update, pass 15-01-2022 for yesterday update
sheet_ranges[Last_column+str(5)].fill =grey_fill
sheet_ranges[Last_column+str(6)].fill =white_smoke_fill
sheet_ranges[Last_column+str(7)].fill =light_green_fill
sheet_ranges[Last_column+str(8)].number_format =numbers.FORMAT_PERCENTAGE
sheet_ranges[Last_column+str(8)]="="+(Last_column+str(7))+"-"+(P_Last_column+str(7))
sheet_ranges[Last_column+str(8)].fill =light_green_fill
sheet_ranges[Last_column+str(9)].number_format =numbers.FORMAT_PERCENTAGE
sheet_ranges[Last_column+str(9)]="=((COUNTIF("+(Last_column+str(start_index_Reg))+":"+(Last_column+str(max_testcases))+","+ '"Failed"'+")"+"/"+Last_column+str(5)+"))"
sheet_ranges[Last_column+str(9)].fill =light_pink_fill
sheet_ranges[Last_column+str(10)].number_format =numbers.FORMAT_PERCENTAGE
sheet_ranges[Last_column+str(10)]="="+(Last_column+str(9))+"-"+(P_Last_column+str(9))
sheet_ranges[Last_column+str(10)].fill =light_pink_fill
sheet_ranges[Last_column+str(11)].number_format =numbers.FORMAT_PERCENTAGE
sheet_ranges[Last_column+str(11)]="=((COUNTIF("+(Last_column+str(start_index_Reg))+":"+(Last_column+str(max_testcases))+","+ '"Error"'+")"+"/"+Last_column+str(5)+"))"
sheet_ranges[Last_column+str(11)].fill =smoke_yellow_fill
sheet_ranges[Last_column+str(12)].number_format =numbers.FORMAT_PERCENTAGE
sheet_ranges[Last_column+str(12)]="="+(Last_column+str(11))+"-"+(P_Last_column+str(11))
sheet_ranges[Last_column+str(12)].fill =smoke_yellow_fill
sheet_ranges[Last_column+str(13)].number_format =numbers.FORMAT_PERCENTAGE
sheet_ranges[Last_column+str(13)]="="+"(("+(Last_column+str(5))+"-"+(Last_column+str(6))+")"+"/"+Last_column+str(5)+")"
sheet_ranges[Last_column+str(13)].fill =light_yellow_fill
#sheet_ranges[Last_column+str(8)]=0.8
time.sleep(5)
Regression.save(Regression_sheet)
Regression.close()
print("Regression sheet is updated")
"""
print("Regression sheet is updated")
if os.path.exists('d:\\Jenkins\\workspace\VWH02_0U0_TA\VWH02_TA_HIL\\VW_SWRT_RegressionCheck.xlsx'):
            os.remove('d:\\Jenkins\\workspace\\VWH02_0U0_TA\\VWH02_TA_HIL\\VW_SWRT_RegressionCheck.xlsx')
shutil.copy2(Regression_sheet, 'd:\\Jenkins\\workspace\\VWH02_0U0_TA\\VWH02_TA_HIL')
"""

"""
Test_cases_list=[]
Test_result_list=[]
start_ind=9
end_ind=165
excel_report = 'Test_report_summary.xlsx'
wb = load_workbook(filename = excel_report)
sheet_ranges = wb['Sheet1'] 
print(sheet_ranges['E9'].value)
for i in range(15,666):
    print(i)
    print(sheet_ranges[('Q'+str(i))].value)
j=0
for i in range(start_ind,end_ind+1):
    Test_cases_list.append(str(sheet_ranges[('B'+str(i))].value))
    Test_result_list.append(str(sheet_ranges[('E'+str(i))].value)) 
print(Test_cases_list)  
print(len(Test_cases_list))
print(len(Test_result_list))
excel_report = 'VW_SWRT_RegressionCheck.xlsx'
wb = load_workbook(filename = excel_report)
sheet_ranges = wb['SWRT Regression Check (auto)'] 
i=0
j=0
test_present=0
start_ind=15
end_ind=665
for j in range(len(Test_cases_list)):
    test_present=0
    for i in range(start_ind,end_ind+1):
        if(Test_cases_list[j] == str(sheet_ranges[('B'+str(i))].value )):
            #print(Test_cases_list[j] + "  "+Test_result_list[j])
            if(Test_result_list[j]=="Success"):
                sheet_ranges[('Q'+str(i))]=Test_result_list[j]
            if(Test_result_list[j]=="Failed"):
                sheet_ranges[('Q'+str(i))]=Test_result_list[j]
            if(Test_result_list[j]=="Error"):
                sheet_ranges[('Q'+str(i))]=Test_result_list[j]
            test_present=1
            break
    if(test_present==0):
        print(" Test case to be added in Regression sheet")
        print(Test_cases_list[j])
print("update done")
wb.save(excel_report)
"""
